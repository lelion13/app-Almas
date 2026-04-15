from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook

from app.services.expense_excel import canonical_payment_method, parse_workbook


def test_canonical_payment_method_trims_and_case() -> None:
    assert canonical_payment_method("  transferencia mercedes  ") == "Transferencia Mercedes"
    assert canonical_payment_method("EFECTIVO") == "Efectivo"
    assert canonical_payment_method("Tarjeta") is None


def test_parse_expense_minimal_headers() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Importe", "Método de Pago"])
    ws.append([Decimal("1500.50"), "Transferencia Irene "])
    buf = BytesIO()
    wb.save(buf)
    result = parse_workbook(buf.getvalue())
    assert not result.row_errors
    assert result.header_row_index == 1
    assert len(result.lines) == 1
    assert result.lines[0].payment_method == "Transferencia Irene"
    assert result.lines[0].amount == Decimal("1500.50")


def test_parse_expense_medio_de_pago_alias() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Medio de pago", "Importe"])
    ws.append(["Transferencia Lea", 200])
    buf = BytesIO()
    wb.save(buf)
    result = parse_workbook(buf.getvalue())
    assert len(result.lines) == 1
    assert result.lines[0].payment_method == "Transferencia Lea"
    assert result.lines[0].amount == Decimal("200")


def test_parse_expense_skips_unknown_method() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Importe", "Método de Pago"])
    ws.append([100, "Efectivo"])
    ws.append([50, "No permitido"])
    buf = BytesIO()
    wb.save(buf)
    result = parse_workbook(buf.getvalue())
    assert len(result.lines) == 1
    assert result.lines[0].amount == Decimal("100")
    assert len(result.row_errors) == 1
    assert "no permitido" in result.row_errors[0].lower()
