from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from app.services.siguefit_excel import (
    H_IMPORTE,
    H_METODO,
    ReportMeta,
    cell_decimal,
    cell_str,
    json_safe_cell_value,
    normalize_group_key,
    normalize_header,
    parse_metadata_rows,
)

H_MEDIO = normalize_header("Medio de pago")

CANONICAL_PAYMENT_METHODS: tuple[str, ...] = (
    "Efectivo",
    "Transferencia Irene",
    "Transferencia Lea",
    "Transferencia Mercedes",
    "Transferencia Raquel",
)


def canonical_payment_method(raw: str | None) -> str | None:
    if not raw:
        return None
    n = normalize_group_key(raw)
    if not n:
        return None
    key = n.casefold()
    for c in CANONICAL_PAYMENT_METHODS:
        if c.casefold() == key:
            return c
    return None


def find_expense_header_row(ws, max_scan: int = 25) -> tuple[int, dict[str, int], str] | None:
    """Returns header row index, column map, and normalized key for method column."""
    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        row_cells = list(ws.iter_rows(min_row=row_idx, max_row=row_idx))[0]
        col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row_cells):
            key = normalize_header(str(cell.value) if cell.value is not None else "")
            if not key:
                continue
            col_map[key] = col_idx
        if H_IMPORTE not in col_map:
            continue
        if H_METODO in col_map:
            return row_idx, col_map, H_METODO
        if H_MEDIO in col_map:
            return row_idx, col_map, H_MEDIO
    return None


@dataclass
class ParsedExpenseLine:
    payment_method: str
    amount: Decimal
    raw_row: dict[str, Any]


@dataclass
class ExpenseParseResult:
    meta: ReportMeta
    lines: list[ParsedExpenseLine]
    header_row_index: int
    rows_skipped: int
    row_errors: list[str]


def parse_workbook(content: bytes) -> ExpenseParseResult:
    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        meta = parse_metadata_rows(ws)
        found = find_expense_header_row(ws)
        if not found:
            return ExpenseParseResult(
                meta=meta,
                lines=[],
                header_row_index=0,
                rows_skipped=0,
                row_errors=[
                    "No se encontró fila de cabecera con Importe y Método de Pago o Medio de pago.",
                ],
            )
        header_row, col_map, method_norm_key = found
        lines: list[ParsedExpenseLine] = []
        errors: list[str] = []
        skipped = 0

        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_cells = list(ws.iter_rows(min_row=row_idx, max_row=row_idx))[0]

            def get_cell(norm_key: str) -> Cell | None:
                if norm_key not in col_map:
                    return None
                i = col_map[norm_key]
                return row_cells[i] if i < len(row_cells) else None

            imp_cell = get_cell(H_IMPORTE)
            amount = cell_decimal(imp_cell)
            if amount is None:
                skipped += 1
                continue

            method_cell = get_cell(method_norm_key)
            raw_method = cell_str(method_cell)
            canonical = canonical_payment_method(raw_method)
            if canonical is None:
                label = raw_method if raw_method else "(vacío)"
                errors.append(f"Fila {row_idx}: medio de pago no permitido: {label!r}.")
                continue

            raw: dict[str, Any] = {}
            for nk, key in [
                (H_IMPORTE, "importe"),
                (method_norm_key, "medio_pago"),
            ]:
                if nk in col_map:
                    i = col_map[nk]
                    v = row_cells[i].value if i < len(row_cells) else None
                    raw[key] = json_safe_cell_value(v)

            lines.append(
                ParsedExpenseLine(
                    payment_method=canonical,
                    amount=amount,
                    raw_row=raw,
                )
            )

        return ExpenseParseResult(
            meta=meta,
            lines=lines,
            header_row_index=header_row,
            rows_skipped=skipped,
            row_errors=errors,
        )
    finally:
        wb.close()
