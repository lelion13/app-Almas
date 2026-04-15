from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


def normalize_header(s: str | None) -> str:
    if s is None:
        return ""
    t = unicodedata.normalize("NFC", str(s)).strip().lower()
    t = re.sub(r"\s+", " ", t)
    return t


def normalize_group_key(s: str | None) -> str:
    if s is None:
        return ""
    t = unicodedata.normalize("NFC", str(s)).strip()
    t = re.sub(r"\s+", " ", t)
    return t


H_FECHA = normalize_header("Fecha")
H_CLIENTE = normalize_header("Cliente")
H_DNI = normalize_header("DNI")
H_CATEGORIA = normalize_header("Categoría de Pago")
H_IMPORTE = normalize_header("Importe")
H_DIVISA = normalize_header("Divisa")
H_METODO = normalize_header("Método de Pago")
H_ACTIVIDAD = normalize_header("Actividad")
H_DETALLE = normalize_header("Detalle")
H_REG = normalize_header("Fecha de Registro")
H_USUARIO = normalize_header("Usuario")

EXCEL_EPOCH = date(1899, 12, 30)


def excel_serial_to_date(serial: Any) -> date | None:
    if serial is None or serial == "":
        return None
    if isinstance(serial, datetime):
        return serial.date()
    if isinstance(serial, date):
        return serial
    try:
        n = float(serial)
    except (TypeError, ValueError):
        return None
    d = EXCEL_EPOCH + timedelta(days=int(n))
    return d


def excel_serial_to_datetime_utc(serial: Any) -> datetime | None:
    if serial is None or serial == "":
        return None
    if isinstance(serial, datetime):
        if serial.tzinfo is None:
            return serial.replace(tzinfo=timezone.utc)
        return serial.astimezone(timezone.utc)
    if isinstance(serial, date):
        return datetime(serial.year, serial.month, serial.day, tzinfo=timezone.utc)
    try:
        n = float(serial)
    except (TypeError, ValueError):
        return None
    whole = int(n)
    frac = n - whole
    d = EXCEL_EPOCH + timedelta(days=whole)
    seconds = int(round(frac * 86400))
    dt = datetime(d.year, d.month, d.day, tzinfo=timezone.utc) + timedelta(seconds=seconds)
    return dt


def cell_str(c: Cell | None) -> str | None:
    if c is None or c.value is None:
        return None
    return str(c.value).strip() if isinstance(c.value, str) else str(c.value)


def cell_decimal(c: Cell | None) -> Decimal | None:
    if c is None or c.value is None or c.value == "":
        return None
    try:
        return Decimal(str(c.value))
    except (InvalidOperation, ValueError):
        return None


def json_safe_cell_value(v: Any) -> Any:
    """Valores aptos para JSON/JSONB (evita datetime no serializable al persistir raw_row)."""
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, str)):
        return v
    if isinstance(v, float):
        return v if v == v else None  # NaN → None
    if isinstance(v, Decimal):
        return str(v)
    if isinstance(v, datetime):
        if v.tzinfo is None:
            return v.isoformat()
        return v.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, time):
        return v.isoformat()
    return str(v)


@dataclass
class ReportMeta:
    source_from: date | None
    source_to: date | None
    activity_filter: str | None


@dataclass
class ParsedLine:
    payment_date: date | None
    client_name: str | None
    dni: str | None
    payment_category: str
    payment_method: str
    amount: Decimal
    currency: str | None
    activity: str | None
    detail: str | None
    registered_at: datetime | None
    registered_by_user: str | None
    raw_row: dict[str, Any]


@dataclass
class ParseResult:
    meta: ReportMeta
    lines: list[ParsedLine]
    header_row_index: int
    rows_skipped: int
    row_errors: list[str]


def parse_metadata_rows(ws) -> ReportMeta:
    source_from = source_to = None
    activity_filter = None
    for row_idx in range(1, min(5, ws.max_row + 1)):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        if not row:
            continue
        a = row[0]
        b = row[1] if len(row) > 1 else None
        c = row[2] if len(row) > 2 else None
        d = row[3] if len(row) > 3 else None
        if normalize_header(str(a) if a is not None else "") == normalize_header("Desde:"):
            source_from = excel_serial_to_date(b)
        if normalize_header(str(c) if c is not None else "") == normalize_header("Hasta:"):
            source_to = excel_serial_to_date(d)
        na = normalize_header(str(a) if a is not None else "")
        if na.startswith("actvidad") or na.startswith("actividad"):
            activity_filter = str(b).strip() if b is not None else None
    return ReportMeta(source_from=source_from, source_to=source_to, activity_filter=activity_filter)


def find_header_row(ws, max_scan: int = 25) -> tuple[int, dict[str, int]] | None:
    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        row_cells = list(ws.iter_rows(min_row=row_idx, max_row=row_idx))[0]
        col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row_cells):
            key = normalize_header(str(cell.value) if cell.value is not None else "")
            if not key:
                continue
            col_map[key] = col_idx
        if H_CATEGORIA in col_map and H_METODO in col_map and H_IMPORTE in col_map:
            return row_idx, col_map
    return None


def parse_workbook(content: bytes) -> ParseResult:
    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        meta = parse_metadata_rows(ws)
        found = find_header_row(ws)
        if not found:
            return ParseResult(
                meta=meta,
                lines=[],
                header_row_index=0,
                rows_skipped=0,
                row_errors=["No se encontró fila de cabecera con Categoría de Pago, Método de Pago e Importe."],
            )
        header_row, col_map = found
        lines: list[ParsedLine] = []
        errors: list[str] = []
        skipped = 0

        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_cells = list(ws.iter_rows(min_row=row_idx, max_row=row_idx))[0]

            def get_cell(norm_key: str) -> Cell | None:
                if norm_key not in col_map:
                    return None
                i = col_map[norm_key]
                return row_cells[i] if i < len(row_cells) else None

            imp_cell = get_cell(H_IMPORTE)
            amount = cell_decimal(imp_cell)
            if amount is None:
                skipped += 1
                continue

            payment_category = normalize_group_key(cell_str(get_cell(H_CATEGORIA))) or ""
            payment_method = normalize_group_key(cell_str(get_cell(H_METODO))) or ""

            raw: dict[str, Any] = {}
            for nk, key in [
                (H_FECHA, "fecha"),
                (H_CLIENTE, "cliente"),
                (H_DNI, "dni"),
                (H_CATEGORIA, "categoria_pago"),
                (H_IMPORTE, "importe"),
                (H_DIVISA, "divisa"),
                (H_METODO, "metodo_pago"),
                (H_ACTIVIDAD, "actividad"),
                (H_DETALLE, "detalle"),
                (H_REG, "fecha_registro"),
                (H_USUARIO, "usuario"),
            ]:
                if nk in col_map:
                    i = col_map[nk]
                    v = row_cells[i].value if i < len(row_cells) else None
                    raw[key] = json_safe_cell_value(v)

            lines.append(
                ParsedLine(
                    payment_date=excel_serial_to_date(get_cell(H_FECHA).value if get_cell(H_FECHA) else None),
                    client_name=normalize_group_key(cell_str(get_cell(H_CLIENTE))) or None,
                    dni=normalize_group_key(cell_str(get_cell(H_DNI))) or None,
                    payment_category=payment_category,
                    payment_method=payment_method,
                    amount=amount,
                    currency=normalize_group_key(cell_str(get_cell(H_DIVISA))) or None,
                    activity=cell_str(get_cell(H_ACTIVIDAD)),
                    detail=cell_str(get_cell(H_DETALLE)),
                    registered_at=excel_serial_to_datetime_utc(
                        get_cell(H_REG).value if get_cell(H_REG) else None
                    ),
                    registered_by_user=normalize_group_key(cell_str(get_cell(H_USUARIO))) or None,
                    raw_row=raw,
                )
            )

        return ParseResult(
            meta=meta,
            lines=lines,
            header_row_index=header_row,
            rows_skipped=skipped,
            row_errors=errors,
        )
    finally:
        wb.close()
